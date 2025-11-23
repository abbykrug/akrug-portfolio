# ------- used for XML parse -------
import xml.etree.ElementTree as ET
# ------- used for GUI ----------
from tkinter import Tk
from tkinter.filedialog import askdirectory
# -------- used for Excel output -------
import os
import openpyxl
from openpyxl.styles import Font

data = [["Organization" , "Reason" , "City, State", "Mission", "CY Revenue", "CY Expenses", "Employees", "Differential", "Start Date", "End Date", "990 year", "EIN", "PY Revenue", "PY Expenses", "BOY Net Assets", "EOY Net Assets", "CY Contributions Recieved", "CY Grants Paid"]]
data_J2 = [["Organization", "Salary(i)", "Incentive(i)", "Other(i)", "Total Cash", "Retirement(i)", "Other Benefits(i)", "Total Benefits", "Total Renumeration(i)", "Percent", "Title", "Name", "EIN", "Salary(ii)", "Incentive(ii)", "Other(ii)", "Retirement(ii)", "Other Benefits(ii)", "Total Renumeration(ii)", "Deferred Comp(i)", "Deferred Comp(ii)" ]]
data_J3 = [["Organization", "Reference", "Explanation"]]
data_J4 = [["Organization", "Total Cash", "Total Renumeration(i)", "Total Renumeration(ii)", "Title"]]
data_J5 = [["Organization", "Name", "Title", "Hrs for Org", "Hrs for Related Org", "Comp from org", "Comp from Related Org", "Estimated Other Comp" ]]

def main():
    p = get_folder()
    loop_over_files(p)
    # parse_xml("./XML_1/Practice1.xml")
    create_output(data, p)

def loop_over_files(p):
    try:
        with os.scandir(p) as entries:
            for entry in entries:
                if entry.is_file() and entry.name.lower().endswith(".xml"):  # Check if the entry is a file
                    parse_xml(entry)
    except FileNotFoundError:
        print(f"Error: Directory '{p}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def get_folder():
    path = askdirectory(title='Select Folder') # shows dialog box and return the path
    return path

def create_output(data, path):
    folder_path = path
    output_filename = 'xml_parser.xlsx'

    # Create the folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path) 
    
    wb = openpyxl.Workbook()
    
    # sheet 1 -- organization information
    sheet_1 = wb.active
    sheet_1.title = "XML_Org"
    for d1 in data:
        sheet_1.append(d1)
    
    # sheet 2 -- Schedule J Part 2
    sheet_2 = wb.create_sheet("Schedule J Part 2")
    for d2 in data_J2:
        sheet_2.append(d2)
    
    # sheet 3 -- Schedule J Part 3
    sheet_3 = wb.create_sheet("Miscellanous")
    for d3 in data_J3:
        sheet_3.append(d3)
    
    # sheet 4 -- Top Two Highest Paid Executives
    sheet_4 = wb.create_sheet("EXEC BY COMP")
    for d4 in data_J4:
        sheet_4.append(d4)
    
    # sheet 5 -- Top Two Highest Paid Executives
    sheet_5 = wb.create_sheet("Section 1 Part 7")
    for d5 in data_J5:
        sheet_5.append(d5)
    

    # ----------------- format all the sheets ----------------
    format_sheet(sheet_1)
    format_sheet(sheet_2)
    format_sheet(sheet_3)
    format_sheet(sheet_4)
    format_sheet(sheet_5)

    # ------------ save worksheet to the folder that was passed in -------------
    full_file_path = os.path.join(folder_path, output_filename)
    wb.save(full_file_path)
    print(full_file_path)
    print(f"Data successfully exported to {output_filename}")

def parse_xml(file_path):
    row = []
    try:
        # 1. Parse the XML file
        tree = ET.parse(file_path)
        root = tree.getroot()

        # IRS namespace
        ns = 'http://www.irs.gov/efile'

        # ---------------------- NAME OF ORGANIZATION ----------------------
        org_name = root.find(f'.//{{{ns}}}BusinessName').find(f'.//{{{ns}}}BusinessNameLine1Txt')
        # ----------------------- City, State, Location -------------------
        city = root.find(f'.//{{{ns}}}USAddress').find(f'.//{{{ns}}}CityNm')
        state = root.find(f'.//{{{ns}}}USAddress').find(f'.//{{{ns}}}StateAbbreviationCd')
        # --------------------- Tax Year Beginning ----------------------
        beg_year = root.find(f'.//{{{ns}}}TaxPeriodBeginDt')
        # -------------------- Tax Year End -------------------------
        end_year = root.find(f'.//{{{ns}}}TaxPeriodEndDt')
        # -------------------- Mission -----------------------------
        mission = root.find(f'.//{{{ns}}}MissionDesc')
        # ----------- Current Year Total Revenue ------------------
        cy_rev = root.find(f'.//{{{ns}}}CYTotalRevenueAmt')
        # -------------- Current Year Total Expenses -----------------
        cy_exp = root.find(f'.//{{{ns}}}CYTotalExpensesAmt')
        # ---------------- Number of Employees ---------------------
        num_em = root.find(f'.//{{{ns}}}TotalEmployeeCnt')
        # ----------------------- EIN --------------------------------
        ein_element = root.find(f'.//{{{ns}}}EIN')
        ein = ein_element.text[:2] + "-" + ein_element.text[2:]
        # ---------- Prior Year Total Revenue ---------------------
        py_rev = root.find(f'.//{{{ns}}}PYTotalRevenueAmt')
        # -------------- Prior Year Total Expenses -------------------
        py_exp = root.find(f'.//{{{ns}}}PYTotalExpensesAmt')
        # ------------- EOY Net Assets ----------------------
        eoy_net = root.find(f'.//{{{ns}}}NetAssetsOrFundBalancesEOYAmt')
        # ------------- BOY Net Assets ----------------------
        boy_net = root.find(f'.//{{{ns}}}NetAssetsOrFundBalancesBOYAmt')
        # ----------- Current Year Contributions Received ------------
        cy_contr = root.find(f'.//{{{ns}}}CYContributionsGrantsAmt')
        # ------------- Current Year Grants Paid --------------------
        cy_grant = root.find(f'.//{{{ns}}}CYGrantsAndSimilarPaidAmt')

        # -------------- add information to list ---------------------
        # organization
        if org_name is not None:
            row.append(org_name.text.title())
        else:
            row.append("n/a")
        
        # space for reason
        row.append(" ")

        # location
        if state is not None and city is not None:
            row.append(city.text.title() + ", " + state.text)
        else:
            row.append("n/a")

        # mission
        if mission is not None:
            row.append(mission.text.capitalize())
        else:
            row.append("n/a")

        # current year revenue
        if cy_rev is not None:
            row.append(trans_num(cy_rev.text))
        else:
            row.append("n/a")

        # current year expenses
        if cy_exp is not None:
            row.append(trans_num(cy_exp.text))
        else:
            row.append("n/a")

        # number of employees
        if num_em is not None:
            row.append(int(num_em.text))
        else:
            row.append("n/a")
        
        # space for differential
        row.append("")

        # beginning year
        if beg_year is not None:
            row.append(format_date(beg_year.text))
        else:
            row.append("n/a")
        
        # end year
        if end_year is not None:
            row.append(format_date(end_year.text))
        else:
            row.append("n/a")

        # space for 990 year
        row.append("")

        # EIN
        if ein_element is not None:
            row.append(ein)
        else:
            row.append("n/a")
        
        # prior year revenue
        if py_rev is not None:
            row.append(trans_num(py_rev.text))
        else:
            row.append("n/a")
        
        # prior year expenses
        if py_exp is not None:
            row.append(trans_num(py_exp.text))
        else:
            row.append("n/a")

        # beginning of year net assets
        if boy_net is not None:
            row.append(trans_num(boy_net.text))
        else:
            row.append("n/a")

        # end of year net assets
        if eoy_net is not None:
            row.append(trans_num(eoy_net.text))
        else:
            row.append("n/a")

        # current year contributions recieved
        if cy_contr is not None:
            row.append(trans_num(cy_contr.text))
        else:
            row.append("n/a")

        # current year grants paid
        if cy_grant is not None:
            row.append(trans_num(cy_grant.text))
        else:
            row.append("n/a")

        data.append(row)

# ------------------------- Schedule J Part II --------------------------------
        scheduleJ = root.find(f'.//{{{ns}}}IRS990ScheduleJ')
        if scheduleJ is None:
            print("Schedule J for " + org_name.text + " not found")
        else:
            info_sort = []
            for s in scheduleJ.findall(f'.//{{{ns}}}RltdOrgOfficerTrstKeyEmplGrp'):
                row_J2 = []
                row_J4 = []
                # --------------- Name and Title --------------------
                sJ_name = s.find(f'{{{ns}}}PersonNm')
                if sJ_name is None:
                    sJ_name = s.find(f'.//{{{ns}}}BusinessName').find(f'.//{{{ns}}}BusinessNameLine1Txt')
                sJ_title = s.find(f'{{{ns}}}TitleTxt')
                # ------------- Base Compensation ------------------
                base_1 = s.find(f'{{{ns}}}BaseCompensationFilingOrgAmt')
                base_2 = s.find(f'{{{ns}}}CompensationBasedOnRltdOrgsAmt')
                # ------------- Bonus & Incentive ------------------
                bonus_1 = s.find(f'{{{ns}}}BonusFilingOrganizationAmount')
                bonus_2 = s.find(f'{{{ns}}}BonusRelatedOrganizationsAmt')
                # ------------ Other Reportable Compensation ------------
                comp_1 = s.find(f'{{{ns}}}OtherCompensationFilingOrgAmt')
                comp_2 = s.find(f'{{{ns}}}OtherCompensationRltdOrgsAmt')
                # ------------- Retirement --------------------
                retire_1 = s.find(f'{{{ns}}}DeferredCompensationFlngOrgAmt')
                retire_2 = s.find(f'{{{ns}}}DeferredCompRltdOrgsAmt')
                # ----------------------- Nontaxable Benefits --------------------
                ben_1 = s.find(f'{{{ns}}}NontaxableBenefitsFilingOrgAmt')
                ben_2 = s.find(f'{{{ns}}}NontaxableBenefitsRltdOrgsAmt')
                # ----------------------- Total Compensation ---------------------
                totc_1 = s.find(f'{{{ns}}}TotalCompensationFilingOrgAmt')
                totc_2 = s.find(f'{{{ns}}}TotalCompensationRltdOrgsAmt')           
                # ---------------------- Deferred Comp ---------------------------
                defc_1 = s.find(f'{{{ns}}}CompReportPrior990FilingOrgAmt')
                defc_2 = s.find(f'{{{ns}}}CompReportPrior990RltdOrgsAmt')
                # -------------- add information to data ---------------------
                row_J2.append(org_name.text.title())
                row_J4.append(org_name.text.title())
                # ----------- Salary(i) ---------------------
                if base_1 is not None:
                    row_J2.append("{:,}".format(int(base_1.text)))
                else:
                    row_J2.append("n/a")
                # ------------- bonus/incentive(i) ------------------------
                if bonus_1 is not None:
                    row_J2.append("{:,}".format(int(bonus_1.text)))
                else:
                    row_J2.append("n/a")
                # ----------- other compensation --------------------
                if comp_1 is not None:
                    row_J2.append("{:,}".format(int(comp_1.text)))
                else:
                    row_J2.append("n/a")
                # ------------ total salary -------------------------
                # total_salary = int(base_1.text) + int(base_2.text) + int(bonus_1.text) + int(bonus_2.text) + int(comp_1.text) + int(comp_2.text)
                # row_J2.append("{:,}".format(total_salary))
                # SPACE FOR TOTAL CASH
                row_J2.append("")
                #row_J4.append("{:,}".format(total_salary))
                row_J4.append("")
                # ------------- Retirement -------------------------
                if retire_1 is not None:
                    row_J2.append("{:,}".format(int(retire_1.text)))
                else:
                    row_J2.append("n/a")
                # ------------- Other --------------------------
                if ben_1 is not None:
                    row_J2.append("{:,}".format(int(ben_1.text)))
                else: 
                    row_J2.append("n/a")
                # ----------------- Total Benefits -------------
                # total_ben = int(retire_1.text) + int(retire_2.text) + int(ben_1.text) + int(ben_2.text)
                #row_J2.append("{:,}".format(total_ben))
                # SPACE FOR TOTAL BENEFITS
                row_J2.append("")
                # ----------------- Renumeration for J2 --------------------
                if totc_1 is not None:
                    row_J2.append("{:,}".format(int(totc_1.text)))
                else: 
                    row_J2.append("n/a")

                # ---------------- Renumeration for J4 ---------------------
                if totc_1 is not None:
                    row_J4.append("{:,}".format(int(totc_1.text)))
                else:
                    row_J4.append("n/a")
            
                if totc_2 is not None:
                    row_J4.append("{:,}".format(int(totc_2.text)))
                else:
                    row_J4.append("n/a")
                
                # SPACE FOR PERCENT 
                row_J2.append("")
                # -------------------

                # -------- name and title ------------
                if sJ_title is not None:
                    row_J2.append(sJ_title.text)
                    row_J4.append(sJ_title.text)
                else:
                    row_J2.append("n/a")
                    row_J4.append("n/a")
                    
                if sJ_name is not None:
                    row_J2.append(sJ_name.text.title())
                else:
                    row_J2.append("n/a")
                # ----------- EIN ----------------
                if ein_element is not None:
                    row_J2.append(ein)
                else:
                    row_J2.append("n/a")
                # ----------- Salary (ii) -----------
                if base_2 is not None:
                    row_J2.append("{:,}".format(int(base_2.text)))
                else:
                    row_J2.append("n/a")
                # ------------ Incentive(ii) ----------
                if bonus_2 is not None:
                    row_J2.append("{:,}".format(int(bonus_2.text)))
                else:
                    row_J2.append("n/a")
                # -------------- Other(ii) -----------
                if comp_2 is not None:
                    row_J2.append("{:,}".format(int(comp_2.text)))
                else:
                    row_J2.append("n/a")
                # -------------- Retirement(ii) --------
                if retire_2 is not None:
                    row_J2.append("{:,}".format(int(retire_2.text)))
                else:
                    row_J2.append("n/a")
                # ------------ Other Benefits(ii) ----------
                if ben_2 is not None:
                    row_J2.append("{:,}".format(int(ben_2.text)))
                else:
                    row_J2.append("n/a")
                # ------------ Total Renumeration (ii) ------
                if totc_2 is not None:
                    row_J2.append("{:,}".format(int(totc_2.text)))
                else:
                    row_J2.append("n/a")
                # ------------ Deferred Comp ------------
                if defc_1 is not None:
                    row_J2.append("{:,}".format(int(defc_1.text)))
                else:
                    row_J2.append("n/a")

                if defc_2 is not None: 
                    row_J2.append("{:,}".format(int(defc_2.text)))
                else:
                    row_J2.append("n/a")
                # ---------------- add to full list ------------
                data_J2.append(row_J2)
                info_sort.append(row_J4)

                # sort this information and add to data_J4
            info_sort.sort(key=lambda sublist: sublist[2])
            for item in reversed(info_sort):
                data_J4.append(item)
            
    # --------------------- Schedule J Part III -----------------------------
            for s in scheduleJ.findall(f'.//{{{ns}}}SupplementalInformationDetail'):
                ref = s.find(f'.//{{{ns}}}FormAndLineReferenceDesc')
                exp = s.find(f'.//{{{ns}}}ExplanationTxt')
                if ref is not None and exp is not None:
                    data_J3.append([org_name.text.title(), ref.text.title(), exp.text.title()])

# -------------------- Part 7, Section 1 ------------------------------
        for ps in root.findall(f'.//{{{ns}}}Form990PartVIISectionAGrp'):
            row_J5 = []
            # --------------- Name and Title --------------------
            ps_name = ps.find(f'{{{ns}}}PersonNm')
            if ps_name is None:
                ps_name = ps.find(f'.//{{{ns}}}BusinessName').find(f'.//{{{ns}}}BusinessNameLine1Txt')
            ps_title = ps.find(f'{{{ns}}}TitleTxt')
            # -------------- Hours for Org ---------------
            hrs_4_org = ps.find(f'{{{ns}}}AverageHoursPerWeekRt')
            # -------------- Hours for Related Org -------------
            hrs_4_oorg = ps.find(f'{{{ns}}}AverageHoursPerWeekRltdOrgRt')
            # -------------- Compensation from Org -------------
            comp_org = ps.find(f'{{{ns}}}ReportableCompFromOrgAmt')
            # -------------- Compensation from Related Org -------
            comp_oorg = ps.find(f'{{{ns}}}ReportableCompFromRltdOrgAmt')
            # -------------- Estimated Compensation ---------------
            est_comp = ps.find(f'{{{ns}}}OtherCompensationAmt')
            # ----------- add information to list -----------------
            # organization 
            row_J5.append(org_name.text.title())

            # name of employee
            if ps_name is not None:
                row_J5.append(ps_name.text.title())
            else:
                row_J5.append("N/A")

            # title of employee
            if ps_title is not None:
                row_J5.append(ps_title.text.title())
            else:
                row_J5.append("N/A")
            
            # hours for organization
            if hrs_4_org is not None:
                row_J5.append(float(hrs_4_org.text))
            else:
                row_J5.append(0.0)

            # hours for related organization
            if hrs_4_oorg is not None:
                row_J5.append(float(hrs_4_oorg.text))
            else:
                row_J5.append(0.0)
            
            # compensation for organization
            if comp_org is not None:
                # c = comp_org.text.format()
                row_J5.append("{:,}".format(int(comp_org.text)))
            else:
                row_J5.append("N/A")

            # compensation for related organization 
            if comp_oorg is not None:
                row_J5.append("{:,}".format(int(comp_oorg.text)))
            else:
                row_J5.append("N/A")

            # estimated other compensation
            if est_comp is not None:
                row_J5.append("{:,}".format(int(est_comp.text)))
            else:
                row_J5.append("N/A")
            
            data_J5.append(row_J5)
        # -------------- END OF THIS ORGANIZATION ------------

        # FORMATTING EXCEL
        data.append([""])
        data_J2.append([""])
        data_J3.append([""])
        data_J4.append([""])
        data_J5.append([""])
    except FileNotFoundError:
        print(f"Error: The file at {file_path} was not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# -------------------- HELPER METHODS -----------------
def trans_num(val):
    v = int(val)//1000
    f_str = "{:,}".format(v)
    return f_str

def format_date(d):
    date = d[5:7] + "/" + d[8:] + "/" + d[:4]
    return date

def format_sheet(sheet):
    header_font = Font(name='Calibri', size=12, bold=True)
    custom_font = Font(name='Calibri', size=12)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = custom_font
    # format headline (this will always be the first row)
    header_row = sheet[1]
    for cell in header_row:
        cell.font = header_font
# -------------------- MAIN ---------------------------
if __name__ == "__main__":
    main()